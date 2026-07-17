"""
Excel 报表导出模块
==================
使用 openpyxl 生成日/月/年度收益报表。
"""
import json
import os
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from models import SettlementParams
from revenue_analytics import (
    discover_daily_results,
    multi_day_revenue_breakdown,
    monthly_revenue_forecast,
    annual_target_tracking,
)


BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.path.join(BASE_DIR, "data", "output")
REPORTS_DIR = os.path.join(OUTPUT_DIR, "reports")


# ============================================================
# 样式工具
# ============================================================

HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True, color="1E293B")
SUBTITLE_FONT = Font(size=11, bold=True, color="64748B")
MONEY_FONT = Font(color="16A34A")
NEGATIVE_FONT = Font(color="DC2626")
THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


def _set_header(ws, row_idx, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _write_row(ws, row_idx, values, money_cols=None, bold=False):
    money_cols = money_cols or []
    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="right" if isinstance(value, (int, float)) else "left")
        if bold:
            cell.font = Font(bold=True)
        if col in money_cols and isinstance(value, (int, float)):
            cell.font = MONEY_FONT if value >= 0 else NEGATIVE_FONT


def _adjust_columns(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)


def _ensure_dir(path: str):
    os.makedirs(os.path.dirname(path), exist_ok=True)


def _write_sample_sheet(wb, sample_rows: List[dict], sheet_name: str = "样本日期明细"):
    """写入样本日期明细工作表。"""
    ws = wb.create_sheet(sheet_name)
    _set_header(ws, 1, ["日期", "是否补全", "净利润(元)", "总收入(元)", "总费用(元)"])
    money_cols = list(range(3, 6))
    for i, r in enumerate(sample_rows, 2):
        _write_row(ws, i, [
            r.get("date", ""),
            "是" if r.get("imputed") else "否",
            r.get("net_profit", 0.0),
            r.get("gross_revenue", 0.0),
            r.get("total_cost", 0.0),
        ], money_cols=money_cols)
    _adjust_columns(ws)


def _write_ranking_sheet(wb, ranking: dict, sheet_name: str = "资产绩效排名"):
    """写入资产绩效排名工作表。"""
    ws = wb.create_sheet(sheet_name)
    _set_header(ws, 1, ["排名", "单元/台区", "额定功率(MW)", "净利润(元)", "单MW净利润(元/MW)", "收益占比(%)"])
    money_cols = [4, 5]
    for i, item in enumerate(ranking.get("ranking", []), 2):
        _write_row(ws, i, [
            item.get("rank", i - 1),
            item.get("name", ""),
            item.get("rated_power_mw", 0.0),
            item.get("net_profit", 0.0),
            item.get("net_profit_per_mw", 0.0),
            item.get("share_pct", 0.0),
        ], money_cols=money_cols)
    _adjust_columns(ws)


# ============================================================
# 日报表导出
# ============================================================

def export_daily_excel(date: str, output_path: str, output_dir: str = OUTPUT_DIR) -> str:
    """导出指定日期的日报表（汇总 + 96 点明细）。"""
    path = os.path.join(output_dir, f"{date}_daily_result.json")
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    summary = data.get("summary", {})
    intervals = data.get("intervals", [])

    wb = Workbook()

    # Sheet 1: 汇总
    ws = wb.active
    ws.title = "汇总"
    ws.cell(row=1, column=1, value=f"储能电站日收益报表 — {date}").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

    info_rows = [
        ("电站", data.get("storage_name", "")),
        ("节点", data.get("node_name", "")),
        ("Kp", summary.get("kp", 1.0)),
        ("日初 SOC", summary.get("initial_soc", 0.5)),
        ("日末 SOC", summary.get("final_soc", 0.5)),
    ]
    for i, (k, v) in enumerate(info_rows, 3):
        _write_row(ws, i, [k, v])

    summary_rows = [
        ("电能量套利收益", summary.get("total_energy_revenue", 0.0)),
        ("调频容量收益", summary.get("total_agc_capacity_revenue", 0.0)),
        ("调频里程收益", summary.get("total_agc_mileage_revenue", 0.0)),
        ("总收入", summary.get("total_gross_revenue", 0.0)),
        ("总费用", summary.get("total_cost", 0.0)),
        ("净利润", summary.get("net_profit", 0.0)),
        ("偏差考核扣款", summary.get("total_deviation_penalty", 0.0)),
        ("总充电量(MWh)", summary.get("total_charge_mwh", 0.0)),
        ("总放电量(MWh)", summary.get("total_discharge_mwh", 0.0)),
    ]
    start = 10
    ws.cell(row=start, column=1, value="收益汇总").font = SUBTITLE_FONT
    _set_header(ws, start + 1, ["项目", "金额(元)", "说明"])
    for i, (name, val) in enumerate(summary_rows, start + 2):
        _write_row(ws, i, [name, val, ""], money_cols=[2])

    # Sheet 2: 分时明细
    ws2 = wb.create_sheet("分时明细")
    headers = [
        "时段", "SOC", "净功率(MW)", "LMP(元/MWh)", "电能量收益", "调频容量收益",
        "调频里程收益", "偏差考核", "输配电费", "政府基金", "辅助分摊",
        "电池衰减", "辅助用电", "净收益",
    ]
    _set_header(ws2, 1, headers)
    money_cols = list(range(5, 15))
    for i, it in enumerate(intervals, 2):
        _write_row(ws2, i, [
            it.get("time", ""),
            it.get("soc", 0.0),
            it.get("net_power_mw", 0.0),
            it.get("lmp", 0.0),
            it.get("energy_revenue", 0.0),
            it.get("agc_capacity_revenue", 0.0),
            it.get("agc_mileage_revenue", 0.0),
            it.get("deviation_penalty", 0.0),
            it.get("transmission_fee", 0.0),
            it.get("gov_fund_fee", 0.0),
            it.get("ancillary_fee", 0.0),
            it.get("battery_degradation_cost", 0.0),
            it.get("self_consumption_cost", 0.0),
            it.get("net_revenue", 0.0),
        ], money_cols=money_cols)

    _adjust_columns(ws)
    _adjust_columns(ws2)
    _ensure_dir(output_path)
    wb.save(output_path)
    return output_path


# ============================================================
# 月度报表导出
# ============================================================

def export_monthly_excel(
    year_month: str,
    output_path: str,
    output_dir: str = OUTPUT_DIR,
    sample_rows: Optional[List[dict]] = None,
    ranking: Optional[dict] = None,
) -> str:
    """导出月度报表（预测 + 日度明细 + 占比 + 样本明细 + 资产排名）。"""
    if sample_rows is None:
        forecast = monthly_revenue_forecast(year_month, output_dir)
        breakdown = multi_day_revenue_breakdown(
            start_date=f"{year_month}-01",
            end_date=f"{year_month}-{forecast['days_in_month']}",
            output_dir=output_dir,
        )
    else:
        from revenue_analytics import monthly_revenue_forecast_by_days
        forecast = monthly_revenue_forecast_by_days(year_month, sample_rows)
        breakdown = {
            "daily": sample_rows,
            "totals": {
                "energy_revenue": sum(r.get("energy_revenue", 0.0) for r in sample_rows),
                "agc_capacity_revenue": sum(r.get("agc_capacity_revenue", 0.0) for r in sample_rows),
                "agc_mileage_revenue": sum(r.get("agc_mileage_revenue", 0.0) for r in sample_rows),
            },
            "component_pct": {},
        }

    wb = Workbook()

    # Sheet 1: 月度预测
    ws = wb.active
    ws.title = "月度预测"
    ws.cell(row=1, column=1, value=f"储能电站月度收益预测 — {year_month}").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

    meta = [
        ("当月天数", forecast["days_in_month"]),
        ("已执行天数", forecast["executed_days"]),
        ("剩余天数", forecast["remaining_days"]),
    ]
    for i, (k, v) in enumerate(meta, 3):
        _write_row(ws, i, [k, v])

    ws.cell(row=7, column=1, value="已执行累计").font = SUBTITLE_FONT
    _set_header(ws, 8, ["项目", "金额(元)"])
    executed = forecast["executed_total"]
    items = [
        ("电能量套利", executed.get("energy_revenue", 0.0)),
        ("调频容量", executed.get("agc_capacity_revenue", 0.0)),
        ("调频里程", executed.get("agc_mileage_revenue", 0.0)),
        ("总收入", executed.get("gross_revenue", 0.0)),
        ("总费用", executed.get("total_cost", 0.0)),
        ("净利润", executed.get("net_profit", 0.0)),
    ]
    for i, (name, val) in enumerate(items, 9):
        _write_row(ws, i, [name, val], money_cols=[2])

    ws.cell(row=17, column=1, value="全月预测").font = SUBTITLE_FONT
    _set_header(ws, 18, ["项目", "点预测(元)", "95% 下限", "95% 上限"])
    fc = forecast["forecast"]
    ci = forecast["confidence_interval"]
    rows = [
        ("净利润", fc.get("net_profit", 0.0), ci["net_profit"]["low"], ci["net_profit"]["high"]),
        ("总收入", fc.get("gross_revenue", 0.0), ci["gross_revenue"]["low"], ci["gross_revenue"]["high"]),
        ("电能量套利", fc.get("energy_revenue", 0.0), "-", "-"),
        ("调频容量", fc.get("agc_capacity_revenue", 0.0), "-", "-"),
        ("调频里程", fc.get("agc_mileage_revenue", 0.0), "-", "-"),
    ]
    for i, row in enumerate(rows, 19):
        _write_row(ws, i, row, money_cols=[2, 3, 4])

    ws.cell(row=26, column=1, value="说明").font = SUBTITLE_FONT
    ws.cell(row=27, column=1, value=forecast.get("note", ""))

    # Sheet 2: 日度明细
    ws2 = wb.create_sheet("日度明细")
    _set_header(ws2, 1, [
        "日期", "电能量收益", "调频容量收益", "调频里程收益",
        "总收入", "总费用", "净利润", "偏差考核", "Kp",
    ])
    money_cols = list(range(2, 9))
    for i, day in enumerate(breakdown["daily"], 2):
        _write_row(ws2, i, [
            day.get("date", ""), day.get("energy_revenue", 0.0), day.get("agc_capacity_revenue", 0.0),
            day.get("agc_mileage_revenue", 0.0), day.get("gross_revenue", 0.0), day.get("total_cost", 0.0),
            day.get("net_profit", 0.0), day.get("total_deviation_penalty", 0.0), day.get("kp", 1.0),
        ], money_cols=money_cols)

    # Sheet 3: 占比分析
    ws3 = wb.create_sheet("占比分析")
    _set_header(ws3, 1, ["收益科目", "金额(元)", "占比(%)"])
    pct = breakdown["component_pct"]
    totals = breakdown["totals"]
    comp_rows = [
        ("电能量套利", totals.get("energy_revenue", 0.0), pct.get("energy", 0.0)),
        ("调频容量", totals.get("agc_capacity_revenue", 0.0), pct.get("agc_capacity", 0.0)),
        ("调频里程", totals.get("agc_mileage_revenue", 0.0), pct.get("agc_mileage", 0.0)),
    ]
    for i, row in enumerate(comp_rows, 2):
        _write_row(ws3, i, row, money_cols=[2])

    # Sheet 4: 样本日期明细
    if sample_rows:
        _write_sample_sheet(wb, sample_rows)

    # Sheet 5: 资产绩效排名
    if ranking:
        _write_ranking_sheet(wb, ranking)

    _adjust_columns(ws)
    _adjust_columns(ws2)
    _adjust_columns(ws3)
    _ensure_dir(output_path)
    wb.save(output_path)
    return output_path


# ============================================================
# 年度报表导出
# ============================================================

def export_annual_excel(
    year: str,
    output_path: str,
    settlement: SettlementParams,
    output_dir: str = OUTPUT_DIR,
    sample_rows: Optional[List[dict]] = None,
    ranking: Optional[dict] = None,
) -> str:
    """导出年度目标追踪报表（支持样本和排名）。"""
    if sample_rows is None:
        tracking = annual_target_tracking(year, settlement, output_dir)
    else:
        from revenue_analytics import annual_target_tracking_by_days
        tracking = annual_target_tracking_by_days(year, sample_rows, settlement)

    wb = Workbook()

    # Sheet 1: 年度目标
    ws = wb.active
    ws.title = "年度目标"
    ws.cell(row=1, column=1, value=f"储能电站年度收益目标追踪 — {year}").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

    _set_header(ws, 3, ["指标", "数值", "单位"])
    rows = [
        ("年度净利润目标", tracking["annual_target_yuan"], "元"),
        ("已执行天数", tracking["executed_days"], "天"),
        ("YTD 净利润", tracking["ytd"]["net_profit"], "元"),
        ("YTD 总收入", tracking["ytd"]["gross_revenue"], "元"),
        ("目标完成进度", tracking["progress_pct"], "%"),
        ("预计全年净利润", tracking["projected_annual"]["net_profit"], "元"),
        ("预计全年总收入", tracking["projected_annual"]["gross_revenue"], "元"),
        ("预计全年缺口/超额", tracking["gap_to_target"], "元"),
    ]
    for i, (name, val, unit) in enumerate(rows, 4):
        display_val = f"{val:.2f}" if isinstance(val, float) else val
        _write_row(ws, i, [name, display_val, unit], money_cols=[2])

    ws.cell(row=14, column=1, value="说明").font = SUBTITLE_FONT
    ws.cell(row=15, column=1, value=tracking.get("note", ""))

    # Sheet 2: 月度预测
    ws2 = wb.create_sheet("月度预测")
    _set_header(ws2, 1, ["月份", "已执行天数", "预测净利润(元)", "预测总收入(元)"])
    for i, m in enumerate(tracking["monthly_forecasts"], 2):
        _write_row(ws2, i, [
            m["year_month"], m["executed_days"],
            m["forecast_net_profit"], m["forecast_gross_revenue"],
        ], money_cols=[3, 4])

    # Sheet 3: 样本日期明细
    if sample_rows:
        _write_sample_sheet(wb, sample_rows)

    # Sheet 4: 资产绩效排名
    if ranking:
        _write_ranking_sheet(wb, ranking)

    _adjust_columns(ws)
    _adjust_columns(ws2)
    _ensure_dir(output_path)
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    settlement = SettlementParams.from_json(
        os.path.join(BASE_DIR, "data", "input", "settlement_params.json")
    )
    print(export_daily_excel("2026-06-30", os.path.join(REPORTS_DIR, "2026-06-30_daily.xlsx")))
    print(export_monthly_excel("2026-06", os.path.join(REPORTS_DIR, "2026-06_monthly.xlsx")))
    print(export_annual_excel("2026", os.path.join(REPORTS_DIR, "2026_annual.xlsx"), settlement))
